"""
05_build_excel_model.py
------------------------
Builds excel/inventory_model.xlsx — a formula-driven, auditable workbook
(openpyxl, no VBA) summarizing the ABC-XYZ / safety-stock / promotion
analysis.

SCALE NOTE: earlier versions of this workbook (575-SKU era) had a
per-SKU "Safety Stock Calculator" tab with one live-formula row per SKU.
At full-dataset scale (166,720 classified SKU-store pairs) a per-SKU
sheet is impractical (huge file, slow to open). This version aggregates
to the 9-cell ABC-XYZ level instead — each of the 9 cells gets one live-
formula row (SS/ROP/HC computed from the cell's aggregate Σstd, Σmean,
count), same formula chain as before, just one row per cell instead of
one row per SKU. This was an explicit user decision (AskUserQuestion,
this session) over "sample a subset of SKUs" or "skip the rebuild".

Run AFTER: scripts/01-04 (needs abc_xyz_matrix, safety_stock_results
in data/inventory.db, and the regenerated data/cleaned/*.csv files)

Usage:
    python scripts/05_build_excel_model.py
"""

import sqlite3
import sys
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, ScatterChart, Reference, Series

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = PROJECT_ROOT / "data" / "inventory.db"
CLEANED_DIR = PROJECT_ROOT / "data" / "cleaned"
EXCEL_PATH = PROJECT_ROOT / "excel" / "inventory_model.xlsx"

LEAD_TIME = 7
HOLDING_COST_RATE = 0.22

CELL_ORDER = ["AX", "AY", "AZ", "BX", "BY", "BZ", "CX", "CY", "CZ"]
CELL_SL = {
    "AX": 0.99, "AY": 0.97, "AZ": 0.95,
    "BX": 0.95, "BY": 0.93, "BZ": 0.90,
    "CX": 0.90, "CY": 0.88, "CZ": 0.85,
}

_stream_handler = logging.StreamHandler(sys.stdout)
_stream_handler.stream = open(sys.stdout.fileno(), mode="w", encoding="utf-8", errors="replace", closefd=False)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                     datefmt="%H:%M:%S", handlers=[_stream_handler])
log = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, size=9, color="666666")
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")


def load_aggregates(conn: sqlite3.Connection) -> dict:
    log.info("Loading aggregates from SQLite...")

    cell_agg = pd.read_sql("""
        SELECT abc_class, xyz_class, cell,
               COUNT(*) AS num_skus,
               SUM(revenue_proxy) AS total_revenue,
               AVG(cv) AS avg_cv,
               SUM(std_daily_demand) AS sum_std_demand,
               SUM(std_daily_demand * unit_cost_proxy) AS sum_std_x_unitcost,
               SUM(mean_daily_demand) AS sum_mean_demand
        FROM abc_xyz_matrix
        GROUP BY abc_class, xyz_class, cell
    """, conn).set_index("cell").loc[CELL_ORDER].reset_index()

    totals = pd.read_sql("""
        SELECT SUM(std_daily_demand) AS sum_std,
               SUM(std_daily_demand * unit_cost_proxy) AS sum_std_cost,
               COUNT(*) AS n
        FROM abc_xyz_matrix
    """, conn).iloc[0]

    abc_xyz_summary = pd.read_csv(CLEANED_DIR / "abc_xyz_summary.csv")
    sensitivity_sl = pd.read_csv(CLEANED_DIR / "sensitivity_service_level.csv")
    promo_summary = pd.read_csv(CLEANED_DIR / "promo_safety_stock_summary.csv")

    log.info(f"   {int(totals['n']):,} classified SKUs, 9 cells aggregated")
    return {
        "cell_agg": cell_agg,
        "totals": totals,
        "abc_xyz_summary": abc_xyz_summary,
        "sensitivity_sl": sensitivity_sl,
        "promo_summary": promo_summary,
    }


def build_config_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Config"
    ws["A1"] = "Inventory Optimization Model — Configuration"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Full 125M-row Corporación Favorita dataset, all 54 stores, 2013-01-01 to "
                "2017-08-15, 166,720 classified SKU-store pairs. ABC-XYZ differentiated "
                "inventory policy, aggregated to the 9-cell level (not per-SKU — see README "
                "for why). All downstream tabs are LIVE formulas off the two inputs below.")
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    ws["A4"] = "Global inputs"
    ws["A4"].font = HEADER_FONT
    ws["A5"] = "Lead time (days)"
    ws["B5"] = LEAD_TIME
    ws["C5"] = "Days between placing and receiving a replenishment order. Sensitivity tested at 3/5/7/10."
    ws["A6"] = "Holding cost rate (annual)"
    ws["B6"] = HOLDING_COST_RATE
    ws["C6"] = "Annual carrying cost as a fraction of unit cost (22%)."

    ws["A8"] = "Unit cost note"
    ws["A8"].font = HEADER_FONT
    ws["A9"] = ("The Favorita dataset has no true unit cost. unit_cost is a proxy = 1.0, "
                "so holding cost is expressed in a unit-cost-proxy currency. Swap in real "
                "costs to get money values.")
    ws["A9"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A9:D9")

    ws["E11"] = "9-Cell Service-Level Policy (VLOOKUP source)"
    ws["E11"].font = HEADER_FONT
    ws["E12"], ws["F12"], ws["G12"] = "Cell", "Service Level", "z-score (NORM.S.INV)"
    for c in ("E12", "F12", "G12"):
        ws[c].font = HEADER_FONT
    for i, cell in enumerate(CELL_ORDER):
        r = 13 + i
        ws[f"E{r}"] = cell
        ws[f"F{r}"] = CELL_SL[cell]
        ws[f"G{r}"] = f"=NORM.S.INV(F{r})"

    for col, w in zip("ABCDEFGH", [26, 10, 60, 4, 6, 8, 10, 10]):
        ws.column_dimensions[col].width = w


def build_abc_xyz_matrix_sheet(wb: Workbook, summary: pd.DataFrame):
    ws = wb.create_sheet("ABC-XYZ Matrix")
    ws["A1"] = "ABC-XYZ Classification Matrix"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("SKU counts and revenue share across the 9 segments (all 54 stores "
                "combined). ABC = revenue tier; XYZ = demand-variability tier.")
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:I2")

    idx = summary.set_index("cell")
    xyz_cols = {"X": "X (stable)", "Y": "Y (moderate)", "Z": "Z (erratic)"}
    abc_rows = {"A": "A (high rev)", "B": "B (mid rev)", "C": "C (low rev)"}

    def grid(top_row, label, value_fn, fmt=None):
        ws[f"A{top_row-1}"] = label
        ws[f"A{top_row-1}"].font = HEADER_FONT
        for j, x in enumerate(["X", "Y", "Z"]):
            ws.cell(row=top_row, column=2 + j, value=xyz_cols[x]).font = HEADER_FONT
        for i, a in enumerate(["A", "B", "C"]):
            ws.cell(row=top_row + 1 + i, column=1, value=abc_rows[a]).font = HEADER_FONT
            for j, x in enumerate(["X", "Y", "Z"]):
                cell_name = a + x
                val = value_fn(cell_name)
                c = ws.cell(row=top_row + 1 + i, column=2 + j, value=val)
                if fmt:
                    c.number_format = fmt

    grid(5, "SKU count per cell", lambda c: int(idx.loc[c, "num_skus"]))
    total_rev = idx["total_revenue"].sum()

    ws["F4"] = "Revenue share % per cell"
    ws["F4"].font = HEADER_FONT
    for j, x in enumerate(["X", "Y", "Z"]):
        ws.cell(row=5, column=7 + j, value=xyz_cols[x]).font = HEADER_FONT
    for i, a in enumerate(["A", "B", "C"]):
        ws.cell(row=6 + i, column=6, value=abc_rows[a]).font = HEADER_FONT
        for j, x in enumerate(["X", "Y", "Z"]):
            cell_name = a + x
            c = ws.cell(row=6 + i, column=7 + j, value=float(idx.loc[cell_name, "total_revenue"]) / total_rev)
            c.number_format = "0.00%"

    ws["A10"] = "Service level policy"
    ws["A10"].font = HEADER_FONT
    for j, x in enumerate(["X", "Y", "Z"]):
        ws.cell(row=11, column=2 + j, value=xyz_cols[x]).font = HEADER_FONT
    for i, a in enumerate(["A", "B", "C"]):
        ws.cell(row=12 + i, column=1, value=abc_rows[a]).font = HEADER_FONT
        for j, x in enumerate(["X", "Y", "Z"]):
            cell_name = a + x
            c = ws.cell(row=12 + i, column=2 + j, value=CELL_SL[cell_name])
            c.number_format = "0%"

    ws["F10"] = "Avg CV per cell"
    ws["F10"].font = HEADER_FONT
    for j, x in enumerate(["X", "Y", "Z"]):
        ws.cell(row=11, column=7 + j, value=xyz_cols[x]).font = HEADER_FONT
    for i, a in enumerate(["A", "B", "C"]):
        ws.cell(row=12 + i, column=6, value=abc_rows[a]).font = HEADER_FONT
        for j, x in enumerate(["X", "Y", "Z"]):
            cell_name = a + x
            c = ws.cell(row=12 + i, column=7 + j, value=round(float(idx.loc[cell_name, "avg_cv"]), 4))

    ws["A16"] = f"Totals: {int(idx['num_skus'].sum()):,} SKU-store pairs, revenue proxy {total_rev:,.0f}"
    ws["A16"].font = NOTE_FONT

    for col, w in zip("ABCDEFGHI", [16, 13, 13, 13, 3, 16, 13, 13, 13]):
        ws.column_dimensions[col].width = w


def build_safety_stock_sheet(wb: Workbook, cell_agg: pd.DataFrame):
    """
    9-cell AGGREGATE safety stock calculator — one live-formula row per
    ABC-XYZ cell (not per SKU; see module docstring for why). Formula
    chain is identical to the old per-SKU version, applied to the cell's
    summed std-dev / mean-demand instead of a single SKU's.
    """
    ws = wb.create_sheet("Safety Stock Calculator")
    ws["A1"] = "Safety Stock Calculator (live formulas, 9-cell aggregate)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("SS = z·Σσ·√LT ; ROP = Σmean·LT + SS ; HC = SS·unit_cost·holding_rate. "
                "z and service level come from the Config policy table via VLOOKUP. "
                "Aggregated to one row per ABC-XYZ cell (166,720 SKUs collapsed to 9 rows) "
                "— edit the policy or the LeadTime/HoldingRate inputs and every row "
                "recalculates.")
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 30

    headers = ["Cell", "SKU Count", "Total Revenue", "Sum Mean Demand", "Sum Std Demand",
               "Unit Cost", "Service Level", "z", "Safety Stock", "Reorder Point", "Holding Cost"]
    for j, h in enumerate(headers):
        c = ws.cell(row=4, column=1 + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    idx = cell_agg.set_index("cell")
    for i, cell in enumerate(CELL_ORDER):
        r = 5 + i
        row = idx.loc[cell]
        ws.cell(row=r, column=1, value=cell)
        ws.cell(row=r, column=2, value=int(row["num_skus"]))
        ws.cell(row=r, column=3, value=float(row["total_revenue"]))
        ws.cell(row=r, column=4, value=float(row["sum_mean_demand"]))
        ws.cell(row=r, column=5, value=float(row["sum_std_demand"]))
        ws.cell(row=r, column=6, value=1.0)  # unit_cost proxy
        ws.cell(row=r, column=7, value=f"=VLOOKUP(A{r},Config!$E$13:$G$21,2,FALSE)")
        ws.cell(row=r, column=8, value=f"=NORM.S.INV(G{r})")
        ws.cell(row=r, column=9, value=f"=H{r}*E{r}*SQRT(Config!$B$5)")
        ws.cell(row=r, column=10, value=f"=D{r}*Config!$B$5+I{r}")
        ws.cell(row=r, column=11, value=f"=I{r}*F{r}*Config!$B$6")

    total_row = 5 + len(CELL_ORDER)
    ws.cell(row=total_row, column=1, value="TOTAL").font = HEADER_FONT
    ws.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row-1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I5:I{total_row-1})")
    ws.cell(row=total_row, column=11, value=f"=SUM(K5:K{total_row-1})")

    for col, w in zip("ABCDEFGHIJK", [6, 10, 16, 16, 14, 9, 12, 8, 13, 13, 13]):
        ws.column_dimensions[col].width = w

    return total_row


def build_cost_tradeoff_sheet(wb: Workbook, sensitivity_sl: pd.DataFrame, ss_total_row: int):
    ws = wb.create_sheet("Cost Tradeoff")
    ws["A1"] = "Cost–Service Level Tradeoff (the money chart)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Uniform service level swept 85%→99.5%. Total SS = z·√LT·Σσ ; Total HC = "
                "z·√LT·holding_rate·Σ(σ·unit_cost). Both are live and react to the Config "
                "inputs. Σσ is pulled from the 9-cell aggregate calculator's grand total "
                "(all 166,720 classified SKUs, regardless of cell — this sweep applies one "
                "uniform SL to everyone). Note the convex curve — the last points of "
                "service level cost the most.")
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 40

    ws["A4"] = "Σ std demand"
    ws["B4"] = f"=SUM('Safety Stock Calculator'!E5:E{ss_total_row-1})"
    ws["A5"] = "Σ (std × unit cost)"
    ws["B5"] = f"=SUMPRODUCT('Safety Stock Calculator'!E5:E{ss_total_row-1},'Safety Stock Calculator'!F5:F{ss_total_row-1})"

    ws["A7"], ws["B7"], ws["C7"], ws["D7"] = "Service Level", "z", "Total Safety Stock", "Total Holding Cost"
    for c in ("A7", "B7", "C7", "D7"):
        ws[c].font = HEADER_FONT

    sls = sensitivity_sl["service_level"].tolist()
    for i, sl in enumerate(sls):
        r = 8 + i
        ws.cell(row=r, column=1, value=float(sl))
        ws.cell(row=r, column=2, value=f"=NORM.S.INV(A{r})")
        ws.cell(row=r, column=3, value=f"=B{r}*SQRT(Config!$B$5)*$B$4")
        ws.cell(row=r, column=4, value=f"=B{r}*SQRT(Config!$B$5)*Config!$B$6*$B$5")

    chart = ScatterChart()
    chart.title = "Total Holding Cost vs. Service Level"
    chart.x_axis.title = "Service Level"
    chart.y_axis.title = "Total Holding Cost"
    last_row = 8 + len(sls) - 1
    xvalues = Reference(ws, min_col=1, min_row=8, max_row=last_row)
    yvalues = Reference(ws, min_col=4, min_row=7, max_row=last_row)
    series = Series(yvalues, xvalues, title_from_data=True)
    series.marker.symbol = "circle"
    series.graphicalProperties.line.noFill = False
    chart.series.append(series)
    ws.add_chart(chart, "F4")

    for col, w in zip("ABCD", [16, 10, 18, 18]):
        ws.column_dimensions[col].width = w


def build_promotion_sheet(wb: Workbook, promo_summary: pd.DataFrame):
    ws = wb.create_sheet("Promotion Adjustments")
    ws["A1"] = "Promotion Adjustments — recommended buffer by family"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("During promotions, demand lifts and the baseline reorder point can be "
                "breached. Extra safety stock needed per family, sized from historical "
                "promo lift, at full-dataset scale (all 54 stores).")
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:G2")

    headers = ["Family", "SKUs", "Lift %", "SS Baseline", "SS Promo", "SS Increase", "HC Increase"]
    for j, h in enumerate(headers):
        c = ws.cell(row=4, column=1 + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    df = promo_summary.sort_values("hc_increase_total", ascending=False).reset_index(drop=True)
    for i, row in df.iterrows():
        r = 5 + i
        ws.cell(row=r, column=1, value=row["family"])
        ws.cell(row=r, column=2, value=int(row["num_skus"]))
        ws.cell(row=r, column=3, value=float(row["lift_pct"]) / 100).number_format = "0.0%"
        ws.cell(row=r, column=4, value=float(row["ss_baseline_total"]))
        ws.cell(row=r, column=5, value=float(row["ss_promo_total"]))
        ws.cell(row=r, column=6, value=float(row["ss_increase_total"]))
        ws.cell(row=r, column=7, value=float(row["hc_increase_total"]))

    last_row = 4 + len(df)
    chart = BarChart()
    chart.title = "HC Increase by Family (top 10)"
    chart.y_axis.title = "Holding cost increase"
    data = Reference(ws, min_col=7, min_row=4, max_row=min(last_row, 14))
    cats = Reference(ws, min_col=1, min_row=5, max_row=min(last_row, 14))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I4")

    for col, w in zip("ABCDEFG", [24, 8, 9, 14, 14, 13, 13]):
        ws.column_dimensions[col].width = w


def build_demand_profile_sheet(wb: Workbook, cell_agg: pd.DataFrame):
    """
    9-cell aggregate demand profile (was per-SKU, 579 rows, in the old
    575-SKU workbook — see module docstring). Scatter of avg mean demand
    vs. avg CV per cell, sized/colored by SKU count, replaces the old
    per-SKU scatter.
    """
    ws = wb.create_sheet("Demand Profile", 1)  # position after Config
    ws["A1"] = "Demand Profile — 9-Cell Aggregate"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Avg demand/CV per ABC-XYZ cell, all 166,720 classified SKU-store pairs "
                "collapsed to 9 rows (was one row per SKU pre-125M-row migration — see "
                "CHANGELOG). CV conditionally colored: green=low variability (X), "
                "red=high (Z).")
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")

    headers = ["Cell", "SKU Count", "Total Revenue", "Avg Mean Demand", "Avg CV", "XYZ Class"]
    for j, h in enumerate(headers):
        c = ws.cell(row=4, column=1 + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    idx = cell_agg.set_index("cell")
    for i, cell in enumerate(CELL_ORDER):
        r = 5 + i
        row = idx.loc[cell]
        n = int(row["num_skus"])
        ws.cell(row=r, column=1, value=cell)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=float(row["total_revenue"]))
        ws.cell(row=r, column=4, value=round(float(row["sum_mean_demand"]) / n, 4))
        ws.cell(row=r, column=5, value=round(float(row["avg_cv"]), 4))
        ws.cell(row=r, column=6, value=cell[1])

    last_row = 4 + len(CELL_ORDER)
    rule = ColorScaleRule(start_type="min", start_color="63BE7B",
                           end_type="max", end_color="F8696B")
    ws.conditional_formatting.add(f"E5:E{last_row}", rule)

    chart = ScatterChart()
    chart.title = "Avg Mean Demand vs. Avg CV, by Cell"
    chart.x_axis.title = "Avg CV"
    chart.y_axis.title = "Avg Mean Demand"
    xvalues = Reference(ws, min_col=5, min_row=5, max_row=last_row)
    yvalues = Reference(ws, min_col=4, min_row=4, max_row=last_row)
    series = Series(yvalues, xvalues, title_from_data=True)
    series.marker.symbol = "diamond"
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    ws.add_chart(chart, "H4")

    for col, w in zip("ABCDEF", [6, 10, 16, 16, 10, 10]):
        ws.column_dimensions[col].width = w


def main():
    log.info("=" * 60)
    log.info("  Inventory Optimization — Excel Model Rebuild (9-cell scale)")
    log.info("=" * 60)

    if not DB_PATH.exists():
        log.error(f"Database not found: {DB_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    try:
        agg = load_aggregates(conn)
    finally:
        conn.close()

    wb = Workbook()
    build_config_sheet(wb)
    build_demand_profile_sheet(wb, agg["cell_agg"])
    build_abc_xyz_matrix_sheet(wb, agg["abc_xyz_summary"])
    ss_total_row = build_safety_stock_sheet(wb, agg["cell_agg"])
    build_cost_tradeoff_sheet(wb, agg["sensitivity_sl"], ss_total_row)
    build_promotion_sheet(wb, agg["promo_summary"])

    EXCEL_PATH.parent.mkdir(exist_ok=True)
    wb.save(EXCEL_PATH)
    log.info(f"Saved {EXCEL_PATH} ({EXCEL_PATH.stat().st_size / 1024:.1f} KB)")
    log.info(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
